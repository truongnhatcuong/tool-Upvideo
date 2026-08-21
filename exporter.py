import os
import csv
from datetime import datetime
from typing import List, Optional


def export_to_excel(
    links: List[str],
    output_path: Optional[str] = None,
    column_name: str = "link",
    export_dir: str = "exports",
) -> str:
    """
    Xuất danh sách link video ra file Excel (.xlsx).
    Nếu chưa cài thư viện openpyxl, tự động fallback về .csv (UTF-8 with BOM mở trực tiếp bằng Excel).

    Args:
        links: Danh sách đường dẫn video.
        output_path: Đường dẫn file xuất (nếu None sẽ tự động đặt tên theo ngày giờ).
        column_name: Tên tiêu đề cột (mặc định 'link').
        export_dir: Thư mục chứa file xuất nếu không chỉ định output_path.

    Returns:
        Đường dẫn file đã được tạo thành công.
    """
    os.makedirs(export_dir, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(export_dir, f"tiktok_links_{timestamp}.xlsx")

    # Đảm bảo thư mục cha của output_path tồn tại
    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    # Thử xuất định dạng .xlsx bằng openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TikTok Links"

        # Style cho Header
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Ghi Header
        ws.cell(row=1, column=1, value=column_name)
        cell = ws.cell(row=1, column=1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.row_dimensions[1].height = 25

        # Ghi từng link
        link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        for row_idx, link in enumerate(links, start=2):
            c = ws.cell(row=row_idx, column=1, value=link)
            c.font = link_font
            c.alignment = Alignment(vertical="center")
            c.border = thin_border
            ws.row_dimensions[row_idx].height = 20

        # Tự động căn độ rộng cột A
        max_len = max([len(str(link)) for link in links] + [len(column_name), 30])
        ws.column_dimensions['A'].width = min(max_len + 5, 100)

        # Lưu file
        wb.save(output_path)
        return os.path.abspath(output_path)

    except ImportError:
        # Fallback về CSV (UTF-8 BOM) nếu máy chưa cài openpyxl
        if not output_path.lower().endswith(".csv"):
            output_path = os.path.splitext(output_path)[0] + ".csv"

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([column_name])
            for link in links:
                writer.writerow([link])

        return os.path.abspath(output_path)


def export_to_csv(
    links: List[str],
    output_path: Optional[str] = None,
    column_name: str = "link",
    export_dir: str = "exports",
) -> str:
    """
    Xuất danh sách link video ra file CSV (UTF-8 BOM hỗ trợ tiếng Việt và Excel).
    """
    os.makedirs(export_dir, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(export_dir, f"tiktok_links_{timestamp}.csv")

    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([column_name])
        for link in links:
            writer.writerow([link])

    return os.path.abspath(output_path)
